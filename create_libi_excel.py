# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

OUTPUT_PATH = r"C:\Users\Avi_Sab\Desktop\ליגת ליבי\LIBI_League_Results_Import.xlsx"

# Colors
ORANGE = "E8620A"
WHITE = "FFFFFF"
GRAY = "888888"
LIGHT_GRAY_ROW = "F2F2F2"
LIGHT_ORANGE_FILL = "FFF3ED"
GREEN_DIFF = "22C55E"
RED_DIFF = "EF4444"
LIGHT_BLUE = "EDF3FF"
LIGHT_ORANGE = "FFF3ED"
LIGHT_YELLOW = "FFFBEB"
WINNER_GREEN = "16A34A"
FINAL_YELLOW = "FEF9C3"
SQL_BG = "F3F4F6"
TOTALS_GRAY = "F2F2F2"


def make_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)


def set_rtl(ws):
    ws.sheet_view.rightToLeft = True


def apply_standings_sheet(ws, title_text, data_rows, note1, note2):
    set_rtl(ws)

    # Row 1: Title
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = title_text
    c.font = Font(name="Arial", size=16, bold=True, color=ORANGE)
    c.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
    ws.row_dimensions[1].height = 28

    # Row 2: Subtitle
    ws.merge_cells("A2:K2")
    c = ws["A2"]
    c.value = "עדכני לאחר סיבוב 8 (21.2.26) | 6 סיבובים נותרו"
    c.font = Font(name="Arial", size=10, italic=True, color=GRAY)
    c.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
    ws.row_dimensions[2].height = 18

    # Row 3: blank
    ws.row_dimensions[3].height = 8

    # Row 4: Header
    headers = ["#", "קבוצה", 'מ"מ', "נצ", "הפ", "עצ", "נק בעד", "נק נגד", "הפרש", "ניכויים", "נקודות"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        c.fill = make_fill(ORANGE)
        c.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
        c.border = Border(bottom=Side(style="thick", color=ORANGE))
    ws.row_dimensions[4].height = 22

    # Data rows
    for i, row_data in enumerate(data_rows):
        row_num = 5 + i
        bg = WHITE if i % 2 == 0 else LIGHT_GRAY_ROW
        for col_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=row_num, column=col_idx, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
            if col_idx == 11:
                # Points column: bold + light orange fill
                c.fill = make_fill(LIGHT_ORANGE_FILL)
                c.font = Font(name="Arial", size=11, bold=True)
            else:
                c.fill = make_fill(bg)
                # Diff column (I=9): colored
                if col_idx == 9:
                    if isinstance(val, (int, float)) and val > 0:
                        c.font = Font(name="Arial", size=11, bold=True, color=GREEN_DIFF)
                    elif isinstance(val, (int, float)) and val < 0:
                        c.font = Font(name="Arial", size=11, bold=True, color=RED_DIFF)
                    else:
                        c.font = Font(name="Arial", size=11)
                else:
                    c.font = Font(name="Arial", size=11)
        ws.row_dimensions[row_num].height = 20

    # Totals/notes separator row
    totals_row = 5 + len(data_rows)
    for col_idx in range(1, 12):
        cell = ws.cell(row=totals_row, column=col_idx)
        cell.fill = make_fill(TOTALS_GRAY)
        cell.border = Border(top=Side(style="medium", color=ORANGE))
    ws.row_dimensions[totals_row].height = 8

    # Note row 1
    note1_row = totals_row + 1
    ws.merge_cells(f"A{note1_row}:K{note1_row}")
    c = ws.cell(row=note1_row, column=1, value=note1)
    c.font = Font(name="Arial", size=8, italic=True, color=GRAY)
    c.alignment = Alignment(horizontal="right", vertical="center", readingOrder=2, wrap_text=True)
    ws.row_dimensions[note1_row].height = 24

    # Note row 2
    note2_row = note1_row + 1
    ws.merge_cells(f"A{note2_row}:K{note2_row}")
    c = ws.cell(row=note2_row, column=1, value=note2)
    c.font = Font(name="Arial", size=8, italic=True, color=GRAY)
    c.alignment = Alignment(horizontal="right", vertical="center", readingOrder=2)
    ws.row_dimensions[note2_row].height = 18

    # Column widths
    widths = {"A": 5, "B": 28, "C": 7, "D": 7, "E": 7, "F": 7,
              "G": 10, "H": 10, "I": 10, "J": 12, "K": 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Freeze row 4
    ws.freeze_panes = "A5"


def create_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ===== SHEET 1: North Standings =====
    ws_north = wb.create_sheet("North Standings")
    north_data = [
        [1, "ידרסל חדרה", 7, 6, 1, 0, 416, 310, 106, -1, 12],
        [2, "חולון", 7, 5, 2, 0, 423, 382, 41, -1, 11],
        [3, "בני נתניה", 7, 4, 3, 0, 374, 369, 5, -1, 10],
        [4, "גוטלמן השרון", 7, 4, 3, 0, 381, 359, 22, -2, 9],
        [5, "בני מוצקין", 6, 3, 3, 0, 362, 325, 37, -2, 7],
        [6, "כ.ע. בת-ים", 7, 2, 3, 2, 287, 343, -56, -2, 5],
        [7, "גלי בת-ים", 7, 0, 5, 2, 212, 367, -155, -1, 4],
    ]
    note1_north = "* ניכויים: בני מוצקין -2 (גמר+מנהלת), חולון -1, חדרה -1, נתניה -1, גוטלמן -2, כ.ע. בת-ים -2, גלי בת-ים -1"
    note2_north = "שיטת הניקוד: נצחון=2 נק׳ | הפסד=1 נק׳ | ספורטק=0 נק׳"
    apply_standings_sheet(ws_north, "ליגת ליבי 2025-2026 — מחוז צפון", north_data, note1_north, note2_north)

    # ===== SHEET 2: South Standings =====
    ws_south = wb.create_sheet("South Standings")
    south_data = [
        [1, 'ראשון "גפן" לציון', 8, 6, 2, 0, 589, 463, 126, 0, 14],
        [2, "אחים קריית משה", 8, 5, 3, 0, 497, 491, 6, 0, 13],
        [3, "קריית מלאכי", 8, 5, 3, 0, 473, 495, -22, 0, 13],
        [4, "אוריה ירושלים", 8, 6, 2, 0, 540, 478, 62, -2, 12],
        [5, "אופק רחובות", 8, 4, 4, 0, 471, 430, 41, 0, 12],
        [6, "אריות קריית גת", 8, 4, 4, 0, 451, 442, 9, 0, 12],
        [7, "אדיס אשדוד", 8, 2, 6, 0, 508, 515, -7, -1, 9],
        [8, "החבר'ה הטובים גדרה", 8, 0, 6, 2, 251, 466, -215, -1, 5],
    ]
    note1_south = "* ניכויים: אוריה ירושלים -2 (מנהלת), אדיס אשדוד -1 (מנהלת), גדרה -1 (מנהלת)"
    note2_south = "שיטת הניקוד: נצחון=2 נק׳ | הפסד=1 נק׳ | ספורטק=0 נק׳"
    apply_standings_sheet(ws_south, "ליגת ליבי 2025-2026 — מחוז דרום", south_data, note1_south, note2_south)

    # ===== SHEET 3: Game Results =====
    ws_games = wb.create_sheet("Game Results")
    set_rtl(ws_games)

    ws_games.merge_cells("A1:K1")
    c = ws_games["A1"]
    c.value = "תוצאות משחקים — סיבובים 1-8"
    c.font = Font(name="Arial", size=16, bold=True, color=ORANGE)
    c.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
    ws_games.row_dimensions[1].height = 28

    ws_games.merge_cells("A2:K2")
    c = ws_games["A2"]
    c.value = "ליגת ליבי 2025-2026"
    c.font = Font(name="Arial", size=10, italic=True, color=GRAY)
    c.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
    ws_games.row_dimensions[2].height = 18

    ws_games.row_dimensions[3].height = 8

    game_headers = ["תאריך", "סיבוב", "מחוז", "קבוצת בית", "תוצאה בית", "תוצאה אורח",
                    "קבוצת אורח", "הפרש", "נקודות בית", "נקודות אורח", "הערה"]
    for col_idx, h in enumerate(game_headers, 1):
        c = ws_games.cell(row=4, column=col_idx, value=h)
        c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        c.fill = make_fill(ORANGE)
        c.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
        c.border = Border(bottom=Side(style="thick", color=ORANGE))
    ws_games.row_dimensions[4].height = 22

    # (date, round, district, home, home_score, away_score, away, diff, pts_home, pts_away, note)
    games = [
        # Round 1
        ("01.11.25", 1, "North", "ידרסל חדרה", 54, 57, "בני נתניה", -3, 1, 2, ""),
        ("01.11.25", 1, "North", "כ.ע. בת-ים", 49, 46, "גלי בת-ים", 3, 2, 1, ""),
        ("01.11.25", 1, "North", "גוטלמן השרון", 51, 59, "חולון", -8, 1, 2, ""),
        ("01.11.25", 1, "South", "אדיס אשדוד", 50, 52, "קריית מלאכי", -2, 1, 2, ""),
        ("01.11.25", 1, "South", "החבר'ה הטובים גדרה", 0, 20, "אריות קריית גת", -20, 0, 2, "טכני לגדרה"),
        ("01.11.25", 1, "South", "אחים קריית משה", 64, 72, 'ראשון "גפן" לציון', -8, 1, 2, ""),
        ("01.11.25", 1, "South", "אוריה ירושלים", 64, 47, "אופק רחובות", 17, 2, 1, ""),
        # Round 2
        ("08.11.25", 2, "North", "בני מוצקין", 43, 50, "ידרסל חדרה", -7, 1, 2, ""),
        ("08.11.25", 2, "North", "גלי בת-ים", 54, 61, "גוטלמן השרון", -7, 1, 2, ""),
        ("08.11.25", 2, "North", "כ.ע. בת-ים", 58, 64, "בני נתניה", -6, 1, 2, ""),
        ("08.11.25", 2, "South", "אוריה ירושלים", 59, 50, "אדיס אשדוד", 9, 2, 1, ""),
        ("08.11.25", 2, "South", "אופק רחובות", 39, 57, "קריית מלאכי", -18, 1, 2, ""),
        ("08.11.25", 2, "South", "החבר'ה הטובים גדרה", 39, 54, 'ראשון "גפן" לציון', -15, 1, 2, ""),
        ("08.11.25", 2, "South", "אריות קריית גת", 67, 68, "אחים קריית משה", -1, 1, 2, ""),
        # Round 3
        ("29.11.25", 3, "North", "בני נתניה", 60, 61, "בני מוצקין", -1, 1, 2, ""),
        ("29.11.25", 3, "North", "גלי בת-ים", 0, 20, "חולון", -20, 0, 2, "טכני לגלי בת-ים"),
        ("29.11.25", 3, "North", "גוטלמן השרון", 20, 0, "כ.ע. בת-ים", 20, 2, 0, "טכני לכח עולה בת-ים"),
        ("29.11.25", 3, "South", "אדיס אשדוד", 74, 42, "החבר'ה הטובים גדרה", 32, 2, 1, ""),
        ("29.11.25", 3, "South", "קריית מלאכי", 65, 64, "אוריה ירושלים", 1, 2, 1, ""),
        ("29.11.25", 3, "South", 'ראשון "גפן" לציון', 64, 53, "אריות קריית גת", 11, 2, 1, ""),
        ("29.11.25", 3, "South", "אופק רחובות", 67, 68, "אחים קריית משה", -1, 1, 2, ""),
        # Round 4
        ("20.12.25", 4, "North", "חולון", 75, 57, "כ.ע. בת-ים", 18, 2, 1, ""),
        ("20.12.25", 4, "North", "בני מוצקין", 53, 63, "גוטלמן השרון", -10, 1, 2, ""),
        ("20.12.25", 4, "North", "ידרסל חדרה", 65, 20, "גלי בת-ים", 45, 2, 1, ""),
        ("20.12.25", 4, "South", "קריית מלאכי", 76, 49, "החבר'ה הטובים גדרה", 27, 2, 1, ""),
        ("20.12.25", 4, "South", "אריות קריית גת", 64, 56, "אדיס אשדוד", 8, 2, 1, ""),
        ("20.12.25", 4, "South", "אחים קריית משה", 78, 80, "אוריה ירושלים", -2, 1, 2, ""),
        ("20.12.25", 4, "South", 'ראשון "גפן" לציון', 70, 54, "אופק רחובות", 16, 2, 1, ""),
        # Round 5
        ("10.01.26", 5, "North", "חולון", 57, 55, "בני מוצקין", 2, 2, 1, ""),
        ("10.01.26", 5, "North", "גוטלמן השרון", 64, 72, "ידרסל חדרה", -8, 1, 2, ""),
        ("10.01.26", 5, "North", "בני נתניה", 20, 0, "גלי בת-ים", 20, 2, 0, "טכני לגלי בת-ים"),
        ("10.01.26", 5, "South", "אדיס אשדוד", 74, 76, "אחים קריית משה", -2, 1, 2, ""),
        ("10.01.26", 5, "South", "אוריה ירושלים", 72, 70, 'ראשון "גפן" לציון', 2, 2, 1, ""),
        ("10.01.26", 5, "South", "אופק רחובות", 72, 27, "החבר'ה הטובים גדרה", 45, 2, 1, ""),
        ("10.01.26", 5, "South", "אריות קריית גת", 60, 56, "קריית מלאכי", 4, 2, 1, ""),
        # Round 6
        ("24.01.26", 6, "North", "חולון", 66, 85, "ידרסל חדרה", -19, 1, 2, ""),
        ("24.01.26", 6, "North", "גוטלמן השרון", 50, 51, "בני נתניה", -1, 1, 2, ""),
        ("24.01.26", 6, "North", "כ.ע. בת-ים", 54, 67, "בני מוצקין", -13, 1, 2, ""),
        ("24.01.26", 6, "South", "קריית מלאכי", 55, 63, "אחים קריית משה", -8, 1, 2, ""),
        ("24.01.26", 6, "South", "אופק רחובות", 76, 51, "אריות קריית גת", 25, 2, 1, ""),
        ("24.01.26", 6, "South", 'ראשון "גפן" לציון', 84, 87, "אדיס אשדוד", -3, 1, 2, ""),
        ("24.01.26", 6, "South", "החבר'ה הטובים גדרה", 45, 81, "אוריה ירושלים", -36, 1, 2, ""),
        # Round 7
        ("25.01.26", 7, "North", "חולון", 76, 62, "בני נתניה", 14, 2, 1, ""),
        ("25.01.26", 7, "North", "ידרסל חדרה", 20, 0, "כ.ע. בת-ים", 20, 2, 0, "טכני לכ.ע. בת-ים"),
        ("25.01.26", 7, "North", "גלי בת-ים", 41, 83, "בני מוצקין", -42, 1, 2, ""),
        ("25.01.26", 7, "South", "אחים קריית משה", 20, 0, "החבר'ה הטובים גדרה", 20, 2, 0, "טכני לגדרה"),
        ("25.01.26", 7, "South", "אדיס אשדוד", 46, 60, "אופק רחובות", -14, 1, 2, ""),
        ("25.01.26", 7, "South", 'ראשון "גפן" לציון', 99, 34, "קריית מלאכי", 65, 2, 1, ""),
        ("25.01.26", 7, "South", "אוריה ירושלים", 73, 67, "אריות קריית גת", 6, 2, 1, ""),
        # Round 8
        ("21.02.26", 8, "North", "חולון", 70, 72, "גוטלמן השרון", -2, 1, 2, ""),
        ("21.02.26", 8, "North", "גלי בת-ים", 51, 69, "כ.ע. בת-ים", -18, 1, 2, ""),
        ("21.02.26", 8, "North", "בני נתניה", 60, 70, "ידרסל חדרה", -10, 1, 2, ""),
        ("21.02.26", 8, "South", "קריית מלאכי", 78, 71, "אדיס אשדוד", 7, 2, 1, ""),
        ("21.02.26", 8, "South", 'ראשון "גפן" לציון', 76, 60, "אחים קריית משה", 16, 2, 1, ""),
        ("21.02.26", 8, "South", "אריות קריית גת", 69, 49, "החבר'ה הטובים גדרה", 20, 2, 1, ""),
        ("21.02.26", 8, "South", "אופק רחובות", 56, 47, "אוריה ירושלים", 9, 2, 1, ""),
    ]

    for i, g in enumerate(games):
        row_num = 5 + i
        date, rnd, district, home, home_sc, away_sc, away, diff, pts_h, pts_a, note = g

        is_forfeit = note.startswith("טכני")
        if is_forfeit:
            bg = LIGHT_YELLOW
        elif district == "North":
            bg = LIGHT_BLUE
        else:
            bg = LIGHT_ORANGE

        row_vals = [date, rnd, district, home, home_sc, away_sc, away, diff, pts_h, pts_a, note]
        for col_idx, val in enumerate(row_vals, 1):
            c = ws_games.cell(row=row_num, column=col_idx, value=val)
            c.fill = make_fill(bg)
            c.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
            if col_idx == 8:  # Diff column
                if isinstance(val, (int, float)) and val > 0:
                    c.font = Font(name="Arial", size=11, bold=True, color=GREEN_DIFF)
                elif isinstance(val, (int, float)) and val < 0:
                    c.font = Font(name="Arial", size=11, bold=True, color=RED_DIFF)
                else:
                    c.font = Font(name="Arial", size=11)
            else:
                c.font = Font(name="Arial", size=11)
            if col_idx in (4, 7):
                c.alignment = Alignment(horizontal="right", vertical="center", readingOrder=2)
        ws_games.row_dimensions[row_num].height = 20

    game_col_widths = {"A": 12, "B": 8, "C": 10, "D": 28, "E": 12, "F": 12,
                       "G": 28, "H": 10, "I": 12, "J": 12, "K": 20}
    for col, width in game_col_widths.items():
        ws_games.column_dimensions[col].width = width

    ws_games.freeze_panes = "A5"
    ws_games.auto_filter.ref = f"A4:K{4 + len(games)}"

    # ===== SHEET 4: Cup Results =====
    ws_cup = wb.create_sheet("Cup Results")
    set_rtl(ws_cup)

    ws_cup.merge_cells("A1:F1")
    c = ws_cup["A1"]
    c.value = "גביע ליגת ליבי 2025-2026"
    c.font = Font(name="Arial", size=16, bold=True, color=ORANGE)
    c.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
    ws_cup.row_dimensions[1].height = 28

    ws_cup.merge_cells("A2:F2")
    c = ws_cup["A2"]
    c.value = "תוצאות עד כה"
    c.font = Font(name="Arial", size=10, italic=True, color=GRAY)
    c.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
    ws_cup.row_dimensions[2].height = 18

    ws_cup.row_dimensions[3].height = 8

    cup_headers = ["שלב", "קבוצת בית", "תוצאה", "קבוצת אורח", "תוצאה", "מנצח"]
    for col_idx, h in enumerate(cup_headers, 1):
        c = ws_cup.cell(row=4, column=col_idx, value=h)
        c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        c.fill = make_fill(ORANGE)
        c.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
        c.border = Border(bottom=Side(style="thick", color=ORANGE))
    ws_cup.row_dimensions[4].height = 22

    # (stage, home, home_sc, away, away_sc, winner, is_final)
    cup_data = [
        ("שמינית גמר (22.11.25)", "גלי בת-ים", 66, "גוטלמן השרון", 75, "גוטלמן השרון", False),
        ("שמינית גמר (22.11.25)", "ידרסל חדרה", 65, "ה.ה. גדרה", 40, "ידרסל חדרה", False),
        ("שמינית גמר (22.11.25)", "אוריה ירושלים", 56, "א.ט. ק. גת", 55, "אוריה ירושלים", False),
        ("שמינית גמר (22.11.25)", "חולון", 61, "אופק רחובות", 54, "חולון", False),
        ("שמינית גמר (22.11.25)", "בני נתניה", 63, "אדיס אשדוד", 49, "בני נתניה", False),
        ("שמינית גמר (22.11.25)", "ק. מלאכי", "20*", "בני מוצקין", "0*", "קריית מלאכי (ספורטק)", False),
        ("שמינית גמר (22.11.25)", 'ראשון "גפן" לציון', "20*", "כ.ע. בת-ים", "0*", "ראשון גפן לציון (ספורטק)", False),
        ("רבע גמר (13.12.25)", 'ראשון "גפן" לציון', 73, "אחים קרית משה", 65, "ראשון גפן לציון", False),
        ("רבע גמר (13.12.25)", "ק. מלאכי", 67, "אוריה ירושלים", 77, "אוריה ירושלים", False),
        ("רבע גמר (13.12.25)", "חולון", 43, "בני נתניה", 64, "בני נתניה", False),
        ("רבע גמר (13.12.25)", "ידרסל חדרה", 38, "גוטלמן השרון", 57, "גוטלמן השרון", False),
        ("חצי גמר (03.01.26)", "אוריה ירושלים", 67, "גוטלמן השרון", 79, "גוטלמן השרון", False),
        ("חצי גמר (03.01.26)", "בני נתניה", 55, 'ראשון "גפן" לציון', 68, "ראשון גפן לציון", False),
        ("גמר (28.3.26)", 'ראשון "גפן" לציון', "TBD", "גוטלמן השרון", "TBD", "טרם נקבע", True),
    ]

    for i, row_data in enumerate(cup_data):
        row_num = 5 + i
        stage, home, home_sc, away, away_sc, winner, is_final = row_data

        if is_final:
            bg = FINAL_YELLOW
        else:
            bg = WHITE if i % 2 == 0 else LIGHT_GRAY_ROW

        vals = [stage, home, home_sc, away, away_sc, winner]
        for col_idx, val in enumerate(vals, 1):
            c = ws_cup.cell(row=row_num, column=col_idx, value=val)
            c.fill = make_fill(bg)
            c.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
            if col_idx == 6 and not is_final:
                c.font = Font(name="Arial", size=11, bold=True, color=WINNER_GREEN)
            elif is_final:
                c.font = Font(name="Arial", size=11, bold=True)
            else:
                c.font = Font(name="Arial", size=11)
        ws_cup.row_dimensions[row_num].height = 20

    cup_col_widths = {"A": 20, "B": 28, "C": 12, "D": 28, "E": 12, "F": 28}
    for col, width in cup_col_widths.items():
        ws_cup.column_dimensions[col].width = width

    ws_cup.freeze_panes = "A5"

    # ===== SHEET 5: SQL Import =====
    ws_sql = wb.create_sheet("SQL Import")
    set_rtl(ws_sql)

    ws_sql["A1"].value = "עדכון תוצאות — הדבק ב-Supabase SQL Editor"
    ws_sql["A1"].font = Font(name="Arial", size=14, bold=True, color=ORANGE)
    ws_sql["A1"].alignment = Alignment(horizontal="right", vertical="center", readingOrder=2)
    ws_sql.row_dimensions[1].height = 24

    ws_sql["A2"].value = "מעדכן את 56 המשחקים שהסתיימו בסיבובים 1-8 לסטטוס Finished עם הניקוד הנכון"
    ws_sql["A2"].font = Font(name="Arial", size=10, italic=True, color=GRAY)
    ws_sql["A2"].alignment = Alignment(horizontal="right", vertical="center", readingOrder=2)
    ws_sql.row_dimensions[2].height = 18

    ws_sql.row_dimensions[3].height = 8

    sql_text = (
        "-- UPDATE GAME RESULTS: Rounds 1-8\n"
        "-- Run this in Supabase SQL Editor to mark all completed games as Finished\n"
        "\n"
        "-- ROUND 1 (2025-11-01)\n"
        "UPDATE games SET home_score=54, away_score=57, status='Finished' WHERE game_date='2025-11-01' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d9\u05d3\u05e8\u05e1\u05dc \u05d7\u05d3\u05e8\u05d4') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d1\u05e0\u05d9 \u05e0\u05ea\u05e0\u05d9\u05d4');\n"
        "UPDATE games SET home_score=49, away_score=46, status='Finished' WHERE game_date='2025-11-01' AND home_team_id=(SELECT id FROM teams WHERE name='\u05db.\u05e2. \u05d1\u05ea-\u05d9\u05dd') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d2\u05dc\u05d9 \u05d1\u05ea-\u05d9\u05dd');\n"
        "UPDATE games SET home_score=51, away_score=59, status='Finished' WHERE game_date='2025-11-01' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d2\u05d5\u05d8\u05dc\u05de\u05df \u05d4\u05e9\u05e8\u05d5\u05df') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d7\u05d5\u05dc\u05d5\u05df');\n"
        "UPDATE games SET home_score=50, away_score=52, status='Finished' WHERE game_date='2025-11-01' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d3\u05d9\u05e1 \u05d0\u05e9\u05d3\u05d5\u05d3') AND away_team_id=(SELECT id FROM teams WHERE name='\u05e7\u05e8\u05d9\u05d9\u05ea \u05de\u05dc\u05d0\u05db\u05d9');\n"
        "UPDATE games SET home_score=0, away_score=20, status='Finished' WHERE game_date='2025-11-01' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d4\u05d7\u05d1\u05e8\u05d4 \u05d4\u05d8\u05d5\u05d1\u05d9\u05dd \u05d2\u05d3\u05e8\u05d4') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05e8\u05d9\u05d5\u05ea \u05e7\u05e8\u05d9\u05d9\u05ea \u05d2\u05ea');\n"
        "UPDATE games SET home_score=64, away_score=72, status='Finished' WHERE game_date='2025-11-01' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d7\u05d9\u05dd \u05e7\u05e8\u05d9\u05d9\u05ea \u05de\u05e9\u05d4') AND away_team_id=(SELECT id FROM teams WHERE name='\u05e8\u05d0\u05e9\u05d5\u05df \u05d2\u05e4\u05df \u05dc\u05e6\u05d9\u05d5\u05df');\n"
        "UPDATE games SET home_score=64, away_score=47, status='Finished' WHERE game_date='2025-11-01' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d5\u05e8\u05d9\u05d4 \u05d9\u05e8\u05d5\u05e9\u05dc\u05d9\u05dd') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d5\u05e4\u05e7 \u05e8\u05d7\u05d5\u05d1\u05d5\u05ea');\n"
        "\n"
        "-- ROUND 2 (2025-11-08)\n"
        "UPDATE games SET home_score=43, away_score=50, status='Finished' WHERE game_date='2025-11-08' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d1\u05e0\u05d9 \u05de\u05d5\u05e6\u05e7\u05d9\u05df') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d9\u05d3\u05e8\u05e1\u05dc \u05d7\u05d3\u05e8\u05d4');\n"
        "UPDATE games SET home_score=54, away_score=61, status='Finished' WHERE game_date='2025-11-08' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d2\u05dc\u05d9 \u05d1\u05ea-\u05d9\u05dd') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d2\u05d5\u05d8\u05dc\u05de\u05df \u05d4\u05e9\u05e8\u05d5\u05df');\n"
        "UPDATE games SET home_score=58, away_score=64, status='Finished' WHERE game_date='2025-11-08' AND home_team_id=(SELECT id FROM teams WHERE name='\u05db.\u05e2. \u05d1\u05ea-\u05d9\u05dd') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d1\u05e0\u05d9 \u05e0\u05ea\u05e0\u05d9\u05d4');\n"
        "UPDATE games SET home_score=59, away_score=50, status='Finished' WHERE game_date='2025-11-08' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d5\u05e8\u05d9\u05d4 \u05d9\u05e8\u05d5\u05e9\u05dc\u05d9\u05dd') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d3\u05d9\u05e1 \u05d0\u05e9\u05d3\u05d5\u05d3');\n"
        "UPDATE games SET home_score=39, away_score=57, status='Finished' WHERE game_date='2025-11-08' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d5\u05e4\u05e7 \u05e8\u05d7\u05d5\u05d1\u05d5\u05ea') AND away_team_id=(SELECT id FROM teams WHERE name='\u05e7\u05e8\u05d9\u05d9\u05ea \u05de\u05dc\u05d0\u05db\u05d9');\n"
        "UPDATE games SET home_score=39, away_score=54, status='Finished' WHERE game_date='2025-11-08' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d4\u05d7\u05d1\u05e8\u05d4 \u05d4\u05d8\u05d5\u05d1\u05d9\u05dd \u05d2\u05d3\u05e8\u05d4') AND away_team_id=(SELECT id FROM teams WHERE name='\u05e8\u05d0\u05e9\u05d5\u05df \u05d2\u05e4\u05df \u05dc\u05e6\u05d9\u05d5\u05df');\n"
        "UPDATE games SET home_score=67, away_score=68, status='Finished' WHERE game_date='2025-11-08' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05e8\u05d9\u05d5\u05ea \u05e7\u05e8\u05d9\u05d9\u05ea \u05d2\u05ea') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d7\u05d9\u05dd \u05e7\u05e8\u05d9\u05d9\u05ea \u05de\u05e9\u05d4');\n"
        "\n"
        "-- ROUND 3 (2025-11-29)\n"
        "UPDATE games SET home_score=60, away_score=61, status='Finished' WHERE game_date='2025-11-29' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d1\u05e0\u05d9 \u05e0\u05ea\u05e0\u05d9\u05d4') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d1\u05e0\u05d9 \u05de\u05d5\u05e6\u05e7\u05d9\u05df');\n"
        "UPDATE games SET home_score=0, away_score=20, status='Finished' WHERE game_date='2025-11-29' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d2\u05dc\u05d9 \u05d1\u05ea-\u05d9\u05dd') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d7\u05d5\u05dc\u05d5\u05df');\n"
        "UPDATE games SET home_score=20, away_score=0, status='Finished' WHERE game_date='2025-11-29' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d2\u05d5\u05d8\u05dc\u05de\u05df \u05d4\u05e9\u05e8\u05d5\u05df') AND away_team_id=(SELECT id FROM teams WHERE name='\u05db.\u05e2. \u05d1\u05ea-\u05d9\u05dd');\n"
        "UPDATE games SET home_score=74, away_score=42, status='Finished' WHERE game_date='2025-11-29' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d3\u05d9\u05e1 \u05d0\u05e9\u05d3\u05d5\u05d3') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d4\u05d7\u05d1\u05e8\u05d4 \u05d4\u05d8\u05d5\u05d1\u05d9\u05dd \u05d2\u05d3\u05e8\u05d4');\n"
        "UPDATE games SET home_score=65, away_score=64, status='Finished' WHERE game_date='2025-11-29' AND home_team_id=(SELECT id FROM teams WHERE name='\u05e7\u05e8\u05d9\u05d9\u05ea \u05de\u05dc\u05d0\u05db\u05d9') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d5\u05e8\u05d9\u05d4 \u05d9\u05e8\u05d5\u05e9\u05dc\u05d9\u05dd');\n"
        "UPDATE games SET home_score=64, away_score=53, status='Finished' WHERE game_date='2025-11-29' AND home_team_id=(SELECT id FROM teams WHERE name='\u05e8\u05d0\u05e9\u05d5\u05df \u05d2\u05e4\u05df \u05dc\u05e6\u05d9\u05d5\u05df') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05e8\u05d9\u05d5\u05ea \u05e7\u05e8\u05d9\u05d9\u05ea \u05d2\u05ea');\n"
        "UPDATE games SET home_score=67, away_score=68, status='Finished' WHERE game_date='2025-11-29' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d5\u05e4\u05e7 \u05e8\u05d7\u05d5\u05d1\u05d5\u05ea') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d7\u05d9\u05dd \u05e7\u05e8\u05d9\u05d9\u05ea \u05de\u05e9\u05d4');\n"
        "\n"
        "-- ROUND 4 (2025-12-20)\n"
        "UPDATE games SET home_score=75, away_score=57, status='Finished' WHERE game_date='2025-12-20' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d7\u05d5\u05dc\u05d5\u05df') AND away_team_id=(SELECT id FROM teams WHERE name='\u05db.\u05e2. \u05d1\u05ea-\u05d9\u05dd');\n"
        "UPDATE games SET home_score=53, away_score=63, status='Finished' WHERE game_date='2025-12-20' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d1\u05e0\u05d9 \u05de\u05d5\u05e6\u05e7\u05d9\u05df') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d2\u05d5\u05d8\u05dc\u05de\u05df \u05d4\u05e9\u05e8\u05d5\u05df');\n"
        "UPDATE games SET home_score=65, away_score=20, status='Finished' WHERE game_date='2025-12-20' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d9\u05d3\u05e8\u05e1\u05dc \u05d7\u05d3\u05e8\u05d4') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d2\u05dc\u05d9 \u05d1\u05ea-\u05d9\u05dd');\n"
        "UPDATE games SET home_score=76, away_score=49, status='Finished' WHERE game_date='2025-12-20' AND home_team_id=(SELECT id FROM teams WHERE name='\u05e7\u05e8\u05d9\u05d9\u05ea \u05de\u05dc\u05d0\u05db\u05d9') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d4\u05d7\u05d1\u05e8\u05d4 \u05d4\u05d8\u05d5\u05d1\u05d9\u05dd \u05d2\u05d3\u05e8\u05d4');\n"
        "UPDATE games SET home_score=64, away_score=56, status='Finished' WHERE game_date='2025-12-20' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05e8\u05d9\u05d5\u05ea \u05e7\u05e8\u05d9\u05d9\u05ea \u05d2\u05ea') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d3\u05d9\u05e1 \u05d0\u05e9\u05d3\u05d5\u05d3');\n"
        "UPDATE games SET home_score=78, away_score=80, status='Finished' WHERE game_date='2025-12-20' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d7\u05d9\u05dd \u05e7\u05e8\u05d9\u05d9\u05ea \u05de\u05e9\u05d4') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d5\u05e8\u05d9\u05d4 \u05d9\u05e8\u05d5\u05e9\u05dc\u05d9\u05dd');\n"
        "UPDATE games SET home_score=70, away_score=54, status='Finished' WHERE game_date='2025-12-20' AND home_team_id=(SELECT id FROM teams WHERE name='\u05e8\u05d0\u05e9\u05d5\u05df \u05d2\u05e4\u05df \u05dc\u05e6\u05d9\u05d5\u05df') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d5\u05e4\u05e7 \u05e8\u05d7\u05d5\u05d1\u05d5\u05ea');\n"
        "\n"
        "-- ROUND 5 (2026-01-10)\n"
        "UPDATE games SET home_score=57, away_score=55, status='Finished' WHERE game_date='2026-01-10' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d7\u05d5\u05dc\u05d5\u05df') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d1\u05e0\u05d9 \u05de\u05d5\u05e6\u05e7\u05d9\u05df');\n"
        "UPDATE games SET home_score=64, away_score=72, status='Finished' WHERE game_date='2026-01-10' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d2\u05d5\u05d8\u05dc\u05de\u05df \u05d4\u05e9\u05e8\u05d5\u05df') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d9\u05d3\u05e8\u05e1\u05dc \u05d7\u05d3\u05e8\u05d4');\n"
        "UPDATE games SET home_score=20, away_score=0, status='Finished' WHERE game_date='2026-01-10' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d1\u05e0\u05d9 \u05e0\u05ea\u05e0\u05d9\u05d4') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d2\u05dc\u05d9 \u05d1\u05ea-\u05d9\u05dd');\n"
        "UPDATE games SET home_score=74, away_score=76, status='Finished' WHERE game_date='2026-01-10' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d3\u05d9\u05e1 \u05d0\u05e9\u05d3\u05d5\u05d3') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d7\u05d9\u05dd \u05e7\u05e8\u05d9\u05d9\u05ea \u05de\u05e9\u05d4');\n"
        "UPDATE games SET home_score=72, away_score=70, status='Finished' WHERE game_date='2026-01-10' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d5\u05e8\u05d9\u05d4 \u05d9\u05e8\u05d5\u05e9\u05dc\u05d9\u05dd') AND away_team_id=(SELECT id FROM teams WHERE name='\u05e8\u05d0\u05e9\u05d5\u05df \u05d2\u05e4\u05df \u05dc\u05e6\u05d9\u05d5\u05df');\n"
        "UPDATE games SET home_score=72, away_score=27, status='Finished' WHERE game_date='2026-01-10' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d5\u05e4\u05e7 \u05e8\u05d7\u05d5\u05d1\u05d5\u05ea') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d4\u05d7\u05d1\u05e8\u05d4 \u05d4\u05d8\u05d5\u05d1\u05d9\u05dd \u05d2\u05d3\u05e8\u05d4');\n"
        "UPDATE games SET home_score=60, away_score=56, status='Finished' WHERE game_date='2026-01-10' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05e8\u05d9\u05d5\u05ea \u05e7\u05e8\u05d9\u05d9\u05ea \u05d2\u05ea') AND away_team_id=(SELECT id FROM teams WHERE name='\u05e7\u05e8\u05d9\u05d9\u05ea \u05de\u05dc\u05d0\u05db\u05d9');\n"
        "\n"
        "-- ROUND 6 (2026-01-24)\n"
        "UPDATE games SET home_score=66, away_score=85, status='Finished' WHERE game_date='2026-01-24' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d7\u05d5\u05dc\u05d5\u05df') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d9\u05d3\u05e8\u05e1\u05dc \u05d7\u05d3\u05e8\u05d4');\n"
        "UPDATE games SET home_score=50, away_score=51, status='Finished' WHERE game_date='2026-01-24' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d2\u05d5\u05d8\u05dc\u05de\u05df \u05d4\u05e9\u05e8\u05d5\u05df') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d1\u05e0\u05d9 \u05e0\u05ea\u05e0\u05d9\u05d4');\n"
        "UPDATE games SET home_score=54, away_score=67, status='Finished' WHERE game_date='2026-01-24' AND home_team_id=(SELECT id FROM teams WHERE name='\u05db.\u05e2. \u05d1\u05ea-\u05d9\u05dd') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d1\u05e0\u05d9 \u05de\u05d5\u05e6\u05e7\u05d9\u05df');\n"
        "UPDATE games SET home_score=55, away_score=63, status='Finished' WHERE game_date='2026-01-24' AND home_team_id=(SELECT id FROM teams WHERE name='\u05e7\u05e8\u05d9\u05d9\u05ea \u05de\u05dc\u05d0\u05db\u05d9') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d7\u05d9\u05dd \u05e7\u05e8\u05d9\u05d9\u05ea \u05de\u05e9\u05d4');\n"
        "UPDATE games SET home_score=76, away_score=51, status='Finished' WHERE game_date='2026-01-24' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d5\u05e4\u05e7 \u05e8\u05d7\u05d5\u05d1\u05d5\u05ea') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05e8\u05d9\u05d5\u05ea \u05e7\u05e8\u05d9\u05d9\u05ea \u05d2\u05ea');\n"
        "UPDATE games SET home_score=84, away_score=87, status='Finished' WHERE game_date='2026-01-24' AND home_team_id=(SELECT id FROM teams WHERE name='\u05e8\u05d0\u05e9\u05d5\u05df \u05d2\u05e4\u05df \u05dc\u05e6\u05d9\u05d5\u05df') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d3\u05d9\u05e1 \u05d0\u05e9\u05d3\u05d5\u05d3');\n"
        "UPDATE games SET home_score=45, away_score=81, status='Finished' WHERE game_date='2026-01-24' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d4\u05d7\u05d1\u05e8\u05d4 \u05d4\u05d8\u05d5\u05d1\u05d9\u05dd \u05d2\u05d3\u05e8\u05d4') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d5\u05e8\u05d9\u05d4 \u05d9\u05e8\u05d5\u05e9\u05dc\u05d9\u05dd');\n"
        "\n"
        "-- ROUND 7 (2026-01-31)\n"
        "UPDATE games SET home_score=76, away_score=62, status='Finished' WHERE game_date='2026-01-31' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d7\u05d5\u05dc\u05d5\u05df') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d1\u05e0\u05d9 \u05e0\u05ea\u05e0\u05d9\u05d4');\n"
        "UPDATE games SET home_score=20, away_score=0, status='Finished' WHERE game_date='2026-01-31' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d9\u05d3\u05e8\u05e1\u05dc \u05d7\u05d3\u05e8\u05d4') AND away_team_id=(SELECT id FROM teams WHERE name='\u05db.\u05e2. \u05d1\u05ea-\u05d9\u05dd');\n"
        "UPDATE games SET home_score=41, away_score=83, status='Finished' WHERE game_date='2026-01-31' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d2\u05dc\u05d9 \u05d1\u05ea-\u05d9\u05dd') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d1\u05e0\u05d9 \u05de\u05d5\u05e6\u05e7\u05d9\u05df');\n"
        "UPDATE games SET home_score=20, away_score=0, status='Finished' WHERE game_date='2026-01-31' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d7\u05d9\u05dd \u05e7\u05e8\u05d9\u05d9\u05ea \u05de\u05e9\u05d4') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d4\u05d7\u05d1\u05e8\u05d4 \u05d4\u05d8\u05d5\u05d1\u05d9\u05dd \u05d2\u05d3\u05e8\u05d4');\n"
        "UPDATE games SET home_score=46, away_score=60, status='Finished' WHERE game_date='2026-01-31' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d3\u05d9\u05e1 \u05d0\u05e9\u05d3\u05d5\u05d3') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d5\u05e4\u05e7 \u05e8\u05d7\u05d5\u05d1\u05d5\u05ea');\n"
        "UPDATE games SET home_score=99, away_score=34, status='Finished' WHERE game_date='2026-01-31' AND home_team_id=(SELECT id FROM teams WHERE name='\u05e8\u05d0\u05e9\u05d5\u05df \u05d2\u05e4\u05df \u05dc\u05e6\u05d9\u05d5\u05df') AND away_team_id=(SELECT id FROM teams WHERE name='\u05e7\u05e8\u05d9\u05d9\u05ea \u05de\u05dc\u05d0\u05db\u05d9');\n"
        "UPDATE games SET home_score=73, away_score=67, status='Finished' WHERE game_date='2026-01-31' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d5\u05e8\u05d9\u05d4 \u05d9\u05e8\u05d5\u05e9\u05dc\u05d9\u05dd') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05e8\u05d9\u05d5\u05ea \u05e7\u05e8\u05d9\u05d9\u05ea \u05d2\u05ea');\n"
        "\n"
        "-- ROUND 8 (2026-02-21)\n"
        "UPDATE games SET home_score=70, away_score=72, status='Finished' WHERE game_date='2026-02-21' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d7\u05d5\u05dc\u05d5\u05df') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d2\u05d5\u05d8\u05dc\u05de\u05df \u05d4\u05e9\u05e8\u05d5\u05df');\n"
        "UPDATE games SET home_score=51, away_score=69, status='Finished' WHERE game_date='2026-02-21' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d2\u05dc\u05d9 \u05d1\u05ea-\u05d9\u05dd') AND away_team_id=(SELECT id FROM teams WHERE name='\u05db.\u05e2. \u05d1\u05ea-\u05d9\u05dd');\n"
        "UPDATE games SET home_score=60, away_score=70, status='Finished' WHERE game_date='2026-02-21' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d1\u05e0\u05d9 \u05e0\u05ea\u05e0\u05d9\u05d4') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d9\u05d3\u05e8\u05e1\u05dc \u05d7\u05d3\u05e8\u05d4');\n"
        "UPDATE games SET home_score=78, away_score=71, status='Finished' WHERE game_date='2026-02-21' AND home_team_id=(SELECT id FROM teams WHERE name='\u05e7\u05e8\u05d9\u05d9\u05ea \u05de\u05dc\u05d0\u05db\u05d9') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d3\u05d9\u05e1 \u05d0\u05e9\u05d3\u05d5\u05d3');\n"
        "UPDATE games SET home_score=76, away_score=60, status='Finished' WHERE game_date='2026-02-21' AND home_team_id=(SELECT id FROM teams WHERE name='\u05e8\u05d0\u05e9\u05d5\u05df \u05d2\u05e4\u05df \u05dc\u05e6\u05d9\u05d5\u05df') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d7\u05d9\u05dd \u05e7\u05e8\u05d9\u05d9\u05ea \u05de\u05e9\u05d4');\n"
        "UPDATE games SET home_score=69, away_score=49, status='Finished' WHERE game_date='2026-02-21' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05e8\u05d9\u05d5\u05ea \u05e7\u05e8\u05d9\u05d9\u05ea \u05d2\u05ea') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d4\u05d7\u05d1\u05e8\u05d4 \u05d4\u05d8\u05d5\u05d1\u05d9\u05dd \u05d2\u05d3\u05e8\u05d4');\n"
        "UPDATE games SET home_score=56, away_score=47, status='Finished' WHERE game_date='2026-02-21' AND home_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d5\u05e4\u05e7 \u05e8\u05d7\u05d5\u05d1\u05d5\u05ea') AND away_team_id=(SELECT id FROM teams WHERE name='\u05d0\u05d5\u05e8\u05d9\u05d4 \u05d9\u05e8\u05d5\u05e9\u05dc\u05d9\u05dd');"
    )

    c = ws_sql["A4"]
    c.value = sql_text
    c.font = Font(name="Courier New", size=9)
    c.fill = make_fill(SQL_BG)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, readingOrder=1)
    ws_sql.row_dimensions[4].height = 800
    ws_sql.column_dimensions["A"].width = 180

    # Save
    wb.save(OUTPUT_PATH)
    print(f"File saved successfully: {OUTPUT_PATH}")


create_workbook()
