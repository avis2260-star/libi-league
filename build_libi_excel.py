import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── helpers ────────────────────────────────────────────────────────────────────
def hex_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def make_font(name="Arial", size=11, bold=False, italic=False, color="000000"):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

ORANGE        = "E8620A"
WHITE         = "FFFFFF"
GRAY          = "6B7280"
GREEN         = "16A34A"
RED           = "DC2626"
ALT_GRAY      = "F2F2F2"
LIGHT_ORANGE  = "FFF3ED"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — North Standings
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "North Standings"

# Row 1 — title
ws1.merge_cells("A1:K1")
c = ws1["A1"]
c.value = "ליגת ליבי 2025-2026 — מחוז צפון"
c.font  = make_font(size=16, bold=True, color=ORANGE)
c.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
ws1.row_dimensions[1].height = 24

# Row 2 — subtitle
ws1["A2"].value = "עדכני לאחר סיבוב 8 | 6 סיבובים נותרו"
ws1["A2"].font  = make_font(size=10, italic=True, color=GRAY)
ws1["A2"].alignment = Alignment(readingOrder=2)

# Row 3 — blank

# Row 4 — headers
north_headers = ["#", "קבוצה", 'מ"מ', "נצ", "הפ", "עצ",
                 "נק בעד", "נק נגד", "הפרש", "ניכויים", "נקודות"]
header_font  = make_font(size=11, bold=True, color=WHITE)
header_fill  = hex_fill(ORANGE)
header_align = Alignment(horizontal="center", vertical="center", readingOrder=2)

for col_i, hdr in enumerate(north_headers, 1):
    cell = ws1.cell(row=4, column=col_i, value=hdr)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = header_align
ws1.row_dimensions[4].height = 18

# Data rows 5-11
north_data = [
    (1, "ידרסל חדרה",       7, 6, 1, 0, 416, 310,  106, -1, 12),
    (2, "חולון",             7, 5, 2, 0, 423, 382,   41, -1, 11),
    (3, "בני נתניה",         7, 4, 3, 0, 374, 369,    5, -1, 10),
    (4, "גוטלמן השרון",      7, 4, 3, 0, 381, 359,   22, -2,  9),
    (5, "בני מוצקין",        6, 3, 3, 0, 362, 325,   37, -2,  7),
    (6, "כ.ע. בת-ים",        7, 2, 3, 2, 287, 343,  -56, -2,  5),
    (7, "גלי בת-ים",         7, 0, 5, 2, 212, 367, -155, -1,  4),
]

for r_idx, row in enumerate(north_data):
    excel_row = r_idx + 5
    fill = hex_fill(WHITE) if r_idx % 2 == 0 else hex_fill(ALT_GRAY)
    for col_i, val in enumerate(row, 1):
        cell = ws1.cell(row=excel_row, column=col_i, value=val)
        cell.fill = fill
        cell.alignment = Alignment(
            horizontal="center" if col_i != 2 else "right",
            vertical="center",
            readingOrder=2
        )
        cell.font = make_font()
        # diff column (I = col 9)
        if col_i == 9:
            color = GREEN if val > 0 else (RED if val < 0 else "000000")
            cell.font = make_font(color=color)
        # points column (K = col 11)
        if col_i == 11:
            cell.font = make_font(bold=True)
            cell.fill = hex_fill(LIGHT_ORANGE)

# Row 12 — note
ws1["A12"].value = "* ניכויים: בני מוצקין -2, חולון -1, חדרה -1, נתניה -1, גוטלמן -2, כ.ע. בת-ים -2, גלי בת-ים -1"
ws1["A12"].font  = make_font(size=8, italic=True, color=GRAY)
ws1["A12"].alignment = Alignment(readingOrder=2)

# Row 13 — scoring
ws1["A13"].value = "שיטת ניקוד: נצחון=2 | הפסד=1 | ספורטק=0"
ws1["A13"].font  = make_font(size=8, italic=True, color=GRAY)
ws1["A13"].alignment = Alignment(readingOrder=2)

# Column widths
col_widths = [5, 28, 7, 7, 7, 7, 10, 10, 10, 11, 11]
for col_i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(col_i)].width = w

# Freeze row 5
ws1.freeze_panes = "A5"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — South Standings
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("South Standings")

ws2.merge_cells("A1:K1")
c = ws2["A1"]
c.value = "ליגת ליבי 2025-2026 — מחוז דרום"
c.font  = make_font(size=16, bold=True, color=ORANGE)
c.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
ws2.row_dimensions[1].height = 24

ws2["A2"].value = "עדכני לאחר סיבוב 8 | 6 סיבובים נותרו"
ws2["A2"].font  = make_font(size=10, italic=True, color=GRAY)
ws2["A2"].alignment = Alignment(readingOrder=2)

for col_i, hdr in enumerate(north_headers, 1):
    cell = ws2.cell(row=4, column=col_i, value=hdr)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = header_align
ws2.row_dimensions[4].height = 18

south_data = [
    (1, 'ראשון "גפן" לציון',       8, 6, 2, 0, 589, 463,  126,  0, 14),
    (2, "אחים קריית משה",           8, 5, 3, 0, 497, 491,    6,  0, 13),
    (3, "קריית מלאכי",              8, 5, 3, 0, 473, 495,  -22,  0, 13),
    (4, "אוריה ירושלים",            8, 6, 2, 0, 540, 478,   62, -2, 12),
    (5, "אופק רחובות",              8, 4, 4, 0, 471, 430,   41,  0, 12),
    (6, "אריות קריית גת",           8, 4, 4, 0, 451, 442,    9,  0, 12),
    (7, "אדיס אשדוד",               8, 2, 6, 0, 508, 515,   -7, -1,  9),
    (8, "החבר'ה הטובים גדרה",       8, 0, 6, 2, 251, 466, -215, -1,  5),
]

for r_idx, row in enumerate(south_data):
    excel_row = r_idx + 5
    fill = hex_fill(WHITE) if r_idx % 2 == 0 else hex_fill(ALT_GRAY)
    for col_i, val in enumerate(row, 1):
        cell = ws2.cell(row=excel_row, column=col_i, value=val)
        cell.fill = fill
        cell.alignment = Alignment(
            horizontal="center" if col_i != 2 else "right",
            vertical="center",
            readingOrder=2
        )
        cell.font = make_font()
        if col_i == 9:
            color = GREEN if val > 0 else (RED if val < 0 else "000000")
            cell.font = make_font(color=color)
        if col_i == 11:
            cell.font = make_font(bold=True)
            cell.fill = hex_fill(LIGHT_ORANGE)

ws2["A13"].value = "* ניכויים: אוריה ירושלים -2, אדיס אשדוד -1, גדרה -1"
ws2["A13"].font  = make_font(size=8, italic=True, color=GRAY)
ws2["A13"].alignment = Alignment(readingOrder=2)

ws2["A14"].value = "שיטת ניקוד: נצחון=2 | הפסד=1 | ספורטק=0"
ws2["A14"].font  = make_font(size=8, italic=True, color=GRAY)
ws2["A14"].alignment = Alignment(readingOrder=2)

for col_i, w in enumerate(col_widths, 1):
    ws2.column_dimensions[get_column_letter(col_i)].width = w

ws2.freeze_panes = "A5"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Game Results
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Game Results")

ws3["A1"].value = "תוצאות משחקים — סיבובים 1-8"
ws3["A1"].font  = make_font(size=14, bold=True, color=ORANGE)
ws3["A1"].alignment = Alignment(readingOrder=2)

ws3["A2"].value = "ליגת ליבי 2025-2026"
ws3["A2"].font  = make_font(size=10, italic=True, color=GRAY)
ws3["A2"].alignment = Alignment(readingOrder=2)

game_headers = ["תאריך", "סיבוב", "מחוז", "קבוצת בית", "בית",
                "אורח", "קבוצת אורח", "הפרש", "נק׳ בית", "נק׳ אורח", "הערה"]

for col_i, hdr in enumerate(game_headers, 1):
    cell = ws3.cell(row=4, column=col_i, value=hdr)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = header_align
ws3.row_dimensions[4].height = 18

# Enable auto-filter on row 4
ws3.auto_filter.ref = "A4:K4"

# games: (date, round, district, home, h_score, a_score, away, diff, h_pts, a_pts, note)
games_raw = [
    ("01.11.25",1,"North","ידרסל חדרה",54,57,"בני נתניה",-3,1,2,""),
    ("01.11.25",1,"North","כ.ע. בת-ים",49,46,"גלי בת-ים",3,2,1,""),
    ("01.11.25",1,"North","גוטלמן השרון",51,59,"חולון",-8,1,2,""),
    ("01.11.25",1,"South","אדיס אשדוד",50,52,"קריית מלאכי",-2,1,2,""),
    ("01.11.25",1,"South","החבר'ה הטובים גדרה",0,20,"אריות קריית גת",-20,0,2,"טכני לגדרה"),
    ("01.11.25",1,"South","אחים קריית משה",64,72,"ראשון גפן לציון",-8,1,2,""),
    ("01.11.25",1,"South","אוריה ירושלים",64,47,"אופק רחובות",17,2,1,""),
    ("08.11.25",2,"North","בני מוצקין",43,50,"ידרסל חדרה",-7,1,2,""),
    ("08.11.25",2,"North","גלי בת-ים",54,61,"גוטלמן השרון",-7,1,2,""),
    ("08.11.25",2,"North","כ.ע. בת-ים",58,64,"בני נתניה",-6,1,2,""),
    ("08.11.25",2,"South","אוריה ירושלים",59,50,"אדיס אשדוד",9,2,1,""),
    ("08.11.25",2,"South","אופק רחובות",39,57,"קריית מלאכי",-18,1,2,""),
    ("08.11.25",2,"South","החבר'ה הטובים גדרה",39,54,"ראשון גפן לציון",-15,1,2,""),
    ("08.11.25",2,"South","אריות קריית גת",67,68,"אחים קריית משה",-1,1,2,""),
    ("29.11.25",3,"North","בני נתניה",60,61,"בני מוצקין",-1,1,2,""),
    ("29.11.25",3,"North","גלי בת-ים",0,20,"חולון",-20,0,2,"טכני לגלי בת-ים"),
    ("29.11.25",3,"North","גוטלמן השרון",20,0,"כ.ע. בת-ים",20,2,0,"טכני לכח עולה"),
    ("29.11.25",3,"South","אדיס אשדוד",74,42,"החבר'ה הטובים גדרה",32,2,1,""),
    ("29.11.25",3,"South","קריית מלאכי",65,64,"אוריה ירושלים",1,2,1,""),
    ("29.11.25",3,"South","ראשון גפן לציון",64,53,"אריות קריית גת",11,2,1,""),
    ("29.11.25",3,"South","אופק רחובות",67,68,"אחים קריית משה",-1,1,2,""),
    ("20.12.25",4,"North","חולון",75,57,"כ.ע. בת-ים",18,2,1,""),
    ("20.12.25",4,"North","בני מוצקין",53,63,"גוטלמן השרון",-10,1,2,""),
    ("20.12.25",4,"North","ידרסל חדרה",65,20,"גלי בת-ים",45,2,1,""),
    ("20.12.25",4,"South","קריית מלאכי",76,49,"החבר'ה הטובים גדרה",27,2,1,""),
    ("20.12.25",4,"South","אריות קריית גת",64,56,"אדיס אשדוד",8,2,1,""),
    ("20.12.25",4,"South","אחים קריית משה",78,80,"אוריה ירושלים",-2,1,2,""),
    ("20.12.25",4,"South","ראשון גפן לציון",70,54,"אופק רחובות",16,2,1,""),
    ("10.01.26",5,"North","חולון",57,55,"בני מוצקין",2,2,1,""),
    ("10.01.26",5,"North","גוטלמן השרון",64,72,"ידרסל חדרה",-8,1,2,""),
    ("10.01.26",5,"North","בני נתניה",20,0,"גלי בת-ים",20,2,0,"טכני לגלי בת-ים"),
    ("10.01.26",5,"South","אדיס אשדוד",74,76,"אחים קריית משה",-2,1,2,""),
    ("10.01.26",5,"South","אוריה ירושלים",72,70,"ראשון גפן לציון",2,2,1,""),
    ("10.01.26",5,"South","אופק רחובות",72,27,"החבר'ה הטובים גדרה",45,2,1,""),
    ("10.01.26",5,"South","אריות קריית גת",60,56,"קריית מלאכי",4,2,1,""),
    ("24.01.26",6,"North","חולון",66,85,"ידרסל חדרה",-19,1,2,""),
    ("24.01.26",6,"North","גוטלמן השרון",50,51,"בני נתניה",-1,1,2,""),
    ("24.01.26",6,"North","כ.ע. בת-ים",54,67,"בני מוצקין",-13,1,2,""),
    ("24.01.26",6,"South","קריית מלאכי",55,63,"אחים קריית משה",-8,1,2,""),
    ("24.01.26",6,"South","אופק רחובות",76,51,"אריות קריית גת",25,2,1,""),
    ("24.01.26",6,"South","ראשון גפן לציון",84,87,"אדיס אשדוד",-3,1,2,""),
    ("24.01.26",6,"South","החבר'ה הטובים גדרה",45,81,"אוריה ירושלים",-36,1,2,""),
    ("31.01.26",7,"North","חולון",76,62,"בני נתניה",14,2,1,""),
    ("31.01.26",7,"North","ידרסל חדרה",20,0,"כ.ע. בת-ים",20,2,0,"טכני לכ.ע. בת-ים"),
    ("31.01.26",7,"North","גלי בת-ים",41,83,"בני מוצקין",-42,1,2,""),
    ("31.01.26",7,"South","אחים קריית משה",20,0,"החבר'ה הטובים גדרה",20,2,0,"טכני לגדרה"),
    ("31.01.26",7,"South","אדיס אשדוד",46,60,"אופק רחובות",-14,1,2,""),
    ("31.01.26",7,"South","ראשון גפן לציון",99,34,"קריית מלאכי",65,2,1,""),
    ("31.01.26",7,"South","אוריה ירושלים",73,67,"אריות קריית גת",6,2,1,""),
    ("21.02.26",8,"North","חולון",70,72,"גוטלמן השרון",-2,1,2,""),
    ("21.02.26",8,"North","גלי בת-ים",51,69,"כ.ע. בת-ים",-18,1,2,""),
    ("21.02.26",8,"North","בני נתניה",60,70,"ידרסל חדרה",-10,1,2,""),
    ("21.02.26",8,"South","קריית מלאכי",78,71,"אדיס אשדוד",7,2,1,""),
    ("21.02.26",8,"South","ראשון גפן לציון",76,60,"אחים קריית משה",16,2,1,""),
    ("21.02.26",8,"South","אריות קריית גת",69,49,"החבר'ה הטובים גדרה",20,2,1,""),
    ("21.02.26",8,"South","אופק רחובות",56,47,"אוריה ירושלים",9,2,1,""),
]

NORTH_FILL   = hex_fill("EDF3FF")
SOUTH_FILL   = hex_fill("FFF3ED")
FORFEIT_FILL = hex_fill("FFFBEB")

def is_forfeit(h, a):
    return (h == 0 and a == 20) or (h == 20 and a == 0)

for r_idx, g in enumerate(games_raw):
    excel_row = r_idx + 5
    date_, rnd, dist, home, hs, as_, away, diff, hp, ap, note = g
    row_vals = [date_, rnd, dist, home, hs, as_, away, diff, hp, ap, note]
    if is_forfeit(hs, as_):
        fill = FORFEIT_FILL
    elif dist == "North":
        fill = NORTH_FILL
    else:
        fill = SOUTH_FILL
    for col_i, val in enumerate(row_vals, 1):
        cell = ws3.cell(row=excel_row, column=col_i, value=val)
        cell.fill = fill
        cell.font = make_font()
        cell.alignment = Alignment(
            horizontal="center" if col_i not in (4, 7, 11) else ("right" if col_i in (4,7) else "left"),
            vertical="center",
            readingOrder=2
        )

game_col_widths = [12, 8, 10, 26, 10, 10, 26, 10, 10, 10, 22]
for col_i, w in enumerate(game_col_widths, 1):
    ws3.column_dimensions[get_column_letter(col_i)].width = w

ws3.freeze_panes = "A5"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Cup Results
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Cup Results")

ws4["A1"].value = "גביע ליגת ליבי 2025-2026"
ws4["A1"].font  = make_font(size=14, bold=True, color=ORANGE)
ws4["A1"].alignment = Alignment(readingOrder=2)

ws4["A2"].value = "טורניר נוקאאוט"
ws4["A2"].font  = make_font(size=10, italic=True, color=GRAY)
ws4["A2"].alignment = Alignment(readingOrder=2)

cup_headers = ["שלב", "קבוצת בית", "תוצאה", "קבוצת אורח", "תוצאה", "מנצח"]
for col_i, hdr in enumerate(cup_headers, 1):
    cell = ws4.cell(row=4, column=col_i, value=hdr)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = header_align

# stage fill colors
R16_FILL   = hex_fill("F0FDF4")
QF_FILL    = hex_fill("DCFCE7")
SF_FILL    = hex_fill("BBF7D0")
FINAL_FILL = hex_fill("FFFBEB")

cup_data = [
    ("שמינית גמר (22.11.25)", "גלי בת-ים",           "66", "גוטלמן השרון",       "75", "גוטלמן השרון",               "R16"),
    ("שמינית גמר (22.11.25)", "ידרסל חדרה",           "65", "ה.ה. גדרה",           "40", "ידרסל חדרה",                 "R16"),
    ("שמינית גמר (22.11.25)", "אוריה ירושלים",        "56", "א.ט. ק. גת",          "55", "אוריה ירושלים",              "R16"),
    ("שמינית גמר (22.11.25)", "חולון",                "61", "אופק רחובות",         "54", "חולון",                      "R16"),
    ("שמינית גמר (22.11.25)", "בני נתניה",            "63", "אדיס אשדוד",          "49", "בני נתניה",                  "R16"),
    ("שמינית גמר (22.11.25)", "קריית מלאכי",          "20", "בני מוצקין",          "0",  "קריית מלאכי (ספורטק)",       "R16"),
    ("שמינית גמר (22.11.25)", "ראשון גפן לציון",      "20", "כ.ע. בת-ים",          "0",  "ראשון גפן לציון (ספורטק)",   "R16"),
    ("רבע גמר (13.12.25)",    "ראשון גפן לציון",      "73", "אחים קריית משה",      "65", "ראשון גפן לציון",            "QF"),
    ("רבע גמר (13.12.25)",    "קריית מלאכי",          "67", "אוריה ירושלים",       "77", "אוריה ירושלים",              "QF"),
    ("רבע גמר (13.12.25)",    "חולון",                "43", "בני נתניה",           "64", "בני נתניה",                  "QF"),
    ("רבע גמר (13.12.25)",    "ידרסל חדרה",           "38", "גוטלמן השרון",        "57", "גוטלמן השרון",               "QF"),
    ("חצי גמר (03.01.26)",    "אוריה ירושלים",        "67", "גוטלמן השרון",        "79", "גוטלמן השרון",               "SF"),
    ("חצי גמר (03.01.26)",    "בני נתניה",            "55", "ראשון גפן לציון",     "68", "ראשון גפן לציון",            "SF"),
    ("גמר (28.3.26)",          "ראשון גפן לציון",      "-",  "גוטלמן השרון",        "-",  "טרם נקבע",                   "Final"),
]

stage_fill_map = {"R16": R16_FILL, "QF": QF_FILL, "SF": SF_FILL, "Final": FINAL_FILL}

for r_idx, row in enumerate(cup_data):
    excel_row = r_idx + 5
    stage_key = row[6]
    fill = stage_fill_map[stage_key]
    for col_i, val in enumerate(row[:6], 1):
        cell = ws4.cell(row=excel_row, column=col_i, value=val)
        cell.fill = fill
        cell.alignment = Alignment(
            horizontal="center" if col_i in (3, 5) else "right",
            vertical="center",
            readingOrder=2
        )
        # winner column (F = col 6)
        if col_i == 6:
            if val == "טרם נקבע":
                cell.font = make_font(bold=True, color=ORANGE)
            else:
                cell.font = make_font(bold=True, color=GREEN)
        else:
            cell.font = make_font()

cup_col_widths = [24, 26, 10, 26, 10, 26]
for col_i, w in enumerate(cup_col_widths, 1):
    ws4.column_dimensions[get_column_letter(col_i)].width = w

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — SQL Import
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("SQL Import")

ws5["A1"].value = "SQL — עדכון תוצאות | הדבק ב-Supabase SQL Editor"
ws5["A1"].font  = make_font(size=13, bold=True, color=ORANGE)
ws5["A1"].alignment = Alignment(readingOrder=2)

ws5["A2"].value = "מעדכן 56 משחקים שהסתיימו בסיבובים 1-8 לסטטוס Finished"
ws5["A2"].font  = make_font(size=10, italic=True, color=GRAY)
ws5["A2"].alignment = Alignment(readingOrder=2)

sql_text = """-- UPDATE GAME RESULTS: Rounds 1-8
-- ROUND 1 (2025-11-01)
UPDATE games SET home_score=54,away_score=57,status='Finished' WHERE game_date='2025-11-01' AND home_team_id=(SELECT id FROM teams WHERE name='ידרסל חדרה') AND away_team_id=(SELECT id FROM teams WHERE name='בני נתניה');
UPDATE games SET home_score=49,away_score=46,status='Finished' WHERE game_date='2025-11-01' AND home_team_id=(SELECT id FROM teams WHERE name='כ.ע. בת-ים') AND away_team_id=(SELECT id FROM teams WHERE name='גלי בת-ים');
UPDATE games SET home_score=51,away_score=59,status='Finished' WHERE game_date='2025-11-01' AND home_team_id=(SELECT id FROM teams WHERE name='גוטלמן השרון') AND away_team_id=(SELECT id FROM teams WHERE name='חולון');
UPDATE games SET home_score=50,away_score=52,status='Finished' WHERE game_date='2025-11-01' AND home_team_id=(SELECT id FROM teams WHERE name='אדיס אשדוד') AND away_team_id=(SELECT id FROM teams WHERE name='קריית מלאכי');
UPDATE games SET home_score=0,away_score=20,status='Finished' WHERE game_date='2025-11-01' AND home_team_id=(SELECT id FROM teams WHERE name='החברה הטובים גדרה') AND away_team_id=(SELECT id FROM teams WHERE name='אריות קריית גת');
UPDATE games SET home_score=64,away_score=72,status='Finished' WHERE game_date='2025-11-01' AND home_team_id=(SELECT id FROM teams WHERE name='אחים קריית משה') AND away_team_id=(SELECT id FROM teams WHERE name='ראשון גפן לציון');
UPDATE games SET home_score=64,away_score=47,status='Finished' WHERE game_date='2025-11-01' AND home_team_id=(SELECT id FROM teams WHERE name='אוריה ירושלים') AND away_team_id=(SELECT id FROM teams WHERE name='אופק רחובות');
-- ROUND 2 (2025-11-08)
UPDATE games SET home_score=43,away_score=50,status='Finished' WHERE game_date='2025-11-08' AND home_team_id=(SELECT id FROM teams WHERE name='בני מוצקין') AND away_team_id=(SELECT id FROM teams WHERE name='ידרסל חדרה');
UPDATE games SET home_score=54,away_score=61,status='Finished' WHERE game_date='2025-11-08' AND home_team_id=(SELECT id FROM teams WHERE name='גלי בת-ים') AND away_team_id=(SELECT id FROM teams WHERE name='גוטלמן השרון');
UPDATE games SET home_score=58,away_score=64,status='Finished' WHERE game_date='2025-11-08' AND home_team_id=(SELECT id FROM teams WHERE name='כ.ע. בת-ים') AND away_team_id=(SELECT id FROM teams WHERE name='בני נתניה');
UPDATE games SET home_score=59,away_score=50,status='Finished' WHERE game_date='2025-11-08' AND home_team_id=(SELECT id FROM teams WHERE name='אוריה ירושלים') AND away_team_id=(SELECT id FROM teams WHERE name='אדיס אשדוד');
UPDATE games SET home_score=39,away_score=57,status='Finished' WHERE game_date='2025-11-08' AND home_team_id=(SELECT id FROM teams WHERE name='אופק רחובות') AND away_team_id=(SELECT id FROM teams WHERE name='קריית מלאכי');
UPDATE games SET home_score=39,away_score=54,status='Finished' WHERE game_date='2025-11-08' AND home_team_id=(SELECT id FROM teams WHERE name='החברה הטובים גדרה') AND away_team_id=(SELECT id FROM teams WHERE name='ראשון גפן לציון');
UPDATE games SET home_score=67,away_score=68,status='Finished' WHERE game_date='2025-11-08' AND home_team_id=(SELECT id FROM teams WHERE name='אריות קריית גת') AND away_team_id=(SELECT id FROM teams WHERE name='אחים קריית משה');
-- ROUND 3 (2025-11-29)
UPDATE games SET home_score=60,away_score=61,status='Finished' WHERE game_date='2025-11-29' AND home_team_id=(SELECT id FROM teams WHERE name='בני נתניה') AND away_team_id=(SELECT id FROM teams WHERE name='בני מוצקין');
UPDATE games SET home_score=0,away_score=20,status='Finished' WHERE game_date='2025-11-29' AND home_team_id=(SELECT id FROM teams WHERE name='גלי בת-ים') AND away_team_id=(SELECT id FROM teams WHERE name='חולון');
UPDATE games SET home_score=20,away_score=0,status='Finished' WHERE game_date='2025-11-29' AND home_team_id=(SELECT id FROM teams WHERE name='גוטלמן השרון') AND away_team_id=(SELECT id FROM teams WHERE name='כ.ע. בת-ים');
UPDATE games SET home_score=74,away_score=42,status='Finished' WHERE game_date='2025-11-29' AND home_team_id=(SELECT id FROM teams WHERE name='אדיס אשדוד') AND away_team_id=(SELECT id FROM teams WHERE name='החברה הטובים גדרה');
UPDATE games SET home_score=65,away_score=64,status='Finished' WHERE game_date='2025-11-29' AND home_team_id=(SELECT id FROM teams WHERE name='קריית מלאכי') AND away_team_id=(SELECT id FROM teams WHERE name='אוריה ירושלים');
UPDATE games SET home_score=64,away_score=53,status='Finished' WHERE game_date='2025-11-29' AND home_team_id=(SELECT id FROM teams WHERE name='ראשון גפן לציון') AND away_team_id=(SELECT id FROM teams WHERE name='אריות קריית גת');
UPDATE games SET home_score=67,away_score=68,status='Finished' WHERE game_date='2025-11-29' AND home_team_id=(SELECT id FROM teams WHERE name='אופק רחובות') AND away_team_id=(SELECT id FROM teams WHERE name='אחים קריית משה');
-- ROUND 4 (2025-12-20)
UPDATE games SET home_score=75,away_score=57,status='Finished' WHERE game_date='2025-12-20' AND home_team_id=(SELECT id FROM teams WHERE name='חולון') AND away_team_id=(SELECT id FROM teams WHERE name='כ.ע. בת-ים');
UPDATE games SET home_score=53,away_score=63,status='Finished' WHERE game_date='2025-12-20' AND home_team_id=(SELECT id FROM teams WHERE name='בני מוצקין') AND away_team_id=(SELECT id FROM teams WHERE name='גוטלמן השרון');
UPDATE games SET home_score=65,away_score=20,status='Finished' WHERE game_date='2025-12-20' AND home_team_id=(SELECT id FROM teams WHERE name='ידרסל חדרה') AND away_team_id=(SELECT id FROM teams WHERE name='גלי בת-ים');
UPDATE games SET home_score=76,away_score=49,status='Finished' WHERE game_date='2025-12-20' AND home_team_id=(SELECT id FROM teams WHERE name='קריית מלאכי') AND away_team_id=(SELECT id FROM teams WHERE name='החברה הטובים גדרה');
UPDATE games SET home_score=64,away_score=56,status='Finished' WHERE game_date='2025-12-20' AND home_team_id=(SELECT id FROM teams WHERE name='אריות קריית גת') AND away_team_id=(SELECT id FROM teams WHERE name='אדיס אשדוד');
UPDATE games SET home_score=78,away_score=80,status='Finished' WHERE game_date='2025-12-20' AND home_team_id=(SELECT id FROM teams WHERE name='אחים קריית משה') AND away_team_id=(SELECT id FROM teams WHERE name='אוריה ירושלים');
UPDATE games SET home_score=70,away_score=54,status='Finished' WHERE game_date='2025-12-20' AND home_team_id=(SELECT id FROM teams WHERE name='ראשון גפן לציון') AND away_team_id=(SELECT id FROM teams WHERE name='אופק רחובות');
-- ROUND 5 (2026-01-10)
UPDATE games SET home_score=57,away_score=55,status='Finished' WHERE game_date='2026-01-10' AND home_team_id=(SELECT id FROM teams WHERE name='חולון') AND away_team_id=(SELECT id FROM teams WHERE name='בני מוצקין');
UPDATE games SET home_score=64,away_score=72,status='Finished' WHERE game_date='2026-01-10' AND home_team_id=(SELECT id FROM teams WHERE name='גוטלמן השרון') AND away_team_id=(SELECT id FROM teams WHERE name='ידרסל חדרה');
UPDATE games SET home_score=20,away_score=0,status='Finished' WHERE game_date='2026-01-10' AND home_team_id=(SELECT id FROM teams WHERE name='בני נתניה') AND away_team_id=(SELECT id FROM teams WHERE name='גלי בת-ים');
UPDATE games SET home_score=74,away_score=76,status='Finished' WHERE game_date='2026-01-10' AND home_team_id=(SELECT id FROM teams WHERE name='אדיס אשדוד') AND away_team_id=(SELECT id FROM teams WHERE name='אחים קריית משה');
UPDATE games SET home_score=72,away_score=70,status='Finished' WHERE game_date='2026-01-10' AND home_team_id=(SELECT id FROM teams WHERE name='אוריה ירושלים') AND away_team_id=(SELECT id FROM teams WHERE name='ראשון גפן לציון');
UPDATE games SET home_score=72,away_score=27,status='Finished' WHERE game_date='2026-01-10' AND home_team_id=(SELECT id FROM teams WHERE name='אופק רחובות') AND away_team_id=(SELECT id FROM teams WHERE name='החברה הטובים גדרה');
UPDATE games SET home_score=60,away_score=56,status='Finished' WHERE game_date='2026-01-10' AND home_team_id=(SELECT id FROM teams WHERE name='אריות קריית גת') AND away_team_id=(SELECT id FROM teams WHERE name='קריית מלאכי');
-- ROUND 6 (2026-01-24)
UPDATE games SET home_score=66,away_score=85,status='Finished' WHERE game_date='2026-01-24' AND home_team_id=(SELECT id FROM teams WHERE name='חולון') AND away_team_id=(SELECT id FROM teams WHERE name='ידרסל חדרה');
UPDATE games SET home_score=50,away_score=51,status='Finished' WHERE game_date='2026-01-24' AND home_team_id=(SELECT id FROM teams WHERE name='גוטלמן השרון') AND away_team_id=(SELECT id FROM teams WHERE name='בני נתניה');
UPDATE games SET home_score=54,away_score=67,status='Finished' WHERE game_date='2026-01-24' AND home_team_id=(SELECT id FROM teams WHERE name='כ.ע. בת-ים') AND away_team_id=(SELECT id FROM teams WHERE name='בני מוצקין');
UPDATE games SET home_score=55,away_score=63,status='Finished' WHERE game_date='2026-01-24' AND home_team_id=(SELECT id FROM teams WHERE name='קריית מלאכי') AND away_team_id=(SELECT id FROM teams WHERE name='אחים קריית משה');
UPDATE games SET home_score=76,away_score=51,status='Finished' WHERE game_date='2026-01-24' AND home_team_id=(SELECT id FROM teams WHERE name='אופק רחובות') AND away_team_id=(SELECT id FROM teams WHERE name='אריות קריית גת');
UPDATE games SET home_score=84,away_score=87,status='Finished' WHERE game_date='2026-01-24' AND home_team_id=(SELECT id FROM teams WHERE name='ראשון גפן לציון') AND away_team_id=(SELECT id FROM teams WHERE name='אדיס אשדוד');
UPDATE games SET home_score=45,away_score=81,status='Finished' WHERE game_date='2026-01-24' AND home_team_id=(SELECT id FROM teams WHERE name='החברה הטובים גדרה') AND away_team_id=(SELECT id FROM teams WHERE name='אוריה ירושלים');
-- ROUND 7 (2026-01-31)
UPDATE games SET home_score=76,away_score=62,status='Finished' WHERE game_date='2026-01-31' AND home_team_id=(SELECT id FROM teams WHERE name='חולון') AND away_team_id=(SELECT id FROM teams WHERE name='בני נתניה');
UPDATE games SET home_score=20,away_score=0,status='Finished' WHERE game_date='2026-01-31' AND home_team_id=(SELECT id FROM teams WHERE name='ידרסל חדרה') AND away_team_id=(SELECT id FROM teams WHERE name='כ.ע. בת-ים');
UPDATE games SET home_score=41,away_score=83,status='Finished' WHERE game_date='2026-01-31' AND home_team_id=(SELECT id FROM teams WHERE name='גלי בת-ים') AND away_team_id=(SELECT id FROM teams WHERE name='בני מוצקין');
UPDATE games SET home_score=20,away_score=0,status='Finished' WHERE game_date='2026-01-31' AND home_team_id=(SELECT id FROM teams WHERE name='אחים קריית משה') AND away_team_id=(SELECT id FROM teams WHERE name='החברה הטובים גדרה');
UPDATE games SET home_score=46,away_score=60,status='Finished' WHERE game_date='2026-01-31' AND home_team_id=(SELECT id FROM teams WHERE name='אדיס אשדוד') AND away_team_id=(SELECT id FROM teams WHERE name='אופק רחובות');
UPDATE games SET home_score=99,away_score=34,status='Finished' WHERE game_date='2026-01-31' AND home_team_id=(SELECT id FROM teams WHERE name='ראשון גפן לציון') AND away_team_id=(SELECT id FROM teams WHERE name='קריית מלאכי');
UPDATE games SET home_score=73,away_score=67,status='Finished' WHERE game_date='2026-01-31' AND home_team_id=(SELECT id FROM teams WHERE name='אוריה ירושלים') AND away_team_id=(SELECT id FROM teams WHERE name='אריות קריית גת');
-- ROUND 8 (2026-02-21)
UPDATE games SET home_score=70,away_score=72,status='Finished' WHERE game_date='2026-02-21' AND home_team_id=(SELECT id FROM teams WHERE name='חולון') AND away_team_id=(SELECT id FROM teams WHERE name='גוטלמן השרון');
UPDATE games SET home_score=51,away_score=69,status='Finished' WHERE game_date='2026-02-21' AND home_team_id=(SELECT id FROM teams WHERE name='גלי בת-ים') AND away_team_id=(SELECT id FROM teams WHERE name='כ.ע. בת-ים');
UPDATE games SET home_score=60,away_score=70,status='Finished' WHERE game_date='2026-02-21' AND home_team_id=(SELECT id FROM teams WHERE name='בני נתניה') AND away_team_id=(SELECT id FROM teams WHERE name='ידרסל חדרה');
UPDATE games SET home_score=78,away_score=71,status='Finished' WHERE game_date='2026-02-21' AND home_team_id=(SELECT id FROM teams WHERE name='קריית מלאכי') AND away_team_id=(SELECT id FROM teams WHERE name='אדיס אשדוד');
UPDATE games SET home_score=76,away_score=60,status='Finished' WHERE game_date='2026-02-21' AND home_team_id=(SELECT id FROM teams WHERE name='ראשון גפן לציון') AND away_team_id=(SELECT id FROM teams WHERE name='אחים קריית משה');
UPDATE games SET home_score=69,away_score=49,status='Finished' WHERE game_date='2026-02-21' AND home_team_id=(SELECT id FROM teams WHERE name='אריות קריית גת') AND away_team_id=(SELECT id FROM teams WHERE name='החברה הטובים גדרה');
UPDATE games SET home_score=56,away_score=47,status='Finished' WHERE game_date='2026-02-21' AND home_team_id=(SELECT id FROM teams WHERE name='אופק רחובות') AND away_team_id=(SELECT id FROM teams WHERE name='אוריה ירושלים');"""

ws5.merge_cells("A4:K4")
sql_cell = ws5["A4"]
sql_cell.value = sql_text
sql_cell.font  = Font(name="Courier New", size=9)
sql_cell.fill  = hex_fill("F3F4F6")
sql_cell.alignment = Alignment(
    wrap_text=True,
    vertical="top",
    readingOrder=1  # LTR for SQL
)
ws5.row_dimensions[4].height = 800
ws5.column_dimensions["A"].width = 120

# ── Save ───────────────────────────────────────────────────────────────────────
output_path = r"C:\Users\Avi_Sab\Desktop\ליגת ליבי\LIBI_League_Results_Import.xlsx"
import os
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"Saved: {output_path}")
