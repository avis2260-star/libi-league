import openpyxl
import sys

wb = openpyxl.load_workbook(
    r'C:\Users\Avi_Sab\Desktop\ליגת ליבי\2025-6\תוצאות וטבלאות לעונת 2025-6 ליגת ליבי(AutoRecovered).xlsx',
    data_only=True
)
print("Sheets:", wb.sheetnames)

# Find the standings sheet
target = None
for s in wb.sheetnames:
    if 'טבלאות' in s or 'tablaot' in s.lower() or 'standings' in s.lower():
        target = s
        break

if target is None:
    # try to find it by index or partial match
    print("Could not find standings sheet. Listing all sheets:")
    for s in wb.sheetnames:
        print(" -", s)
    sys.exit(1)

print(f"\nReading sheet: '{target}'")
ws = wb[target]

print(f"\nDimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}\n")

print("=== ALL CELLS (row x col : value) ===")
for row in ws.iter_rows():
    for cell in row:
        if cell.value is not None:
            print(f"  R{cell.row}C{cell.column} ({cell.coordinate}): {repr(cell.value)}")
