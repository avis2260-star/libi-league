import openpyxl
import sys

out = open(r'C:\Users\Avi_Sab\New folder\standings_output.txt', 'w', encoding='utf-8')

def p(s=''):
    out.write(str(s) + '\n')
    out.flush()

try:
    wb = openpyxl.load_workbook(
        r'C:\Users\Avi_Sab\Desktop\ליגת ליבי\2025-6\תוצאות וטבלאות לעונת 2025-6 ליגת ליבי(AutoRecovered).xlsx',
        data_only=True
    )
    p("Sheets: " + str(wb.sheetnames))

    target = None
    for s in wb.sheetnames:
        if 'טבלאות' in s:
            target = s
            break
    if target is None:
        for s in wb.sheetnames:
            p(" - " + s)
        out.close()
        sys.exit(1)

    p(f"Reading sheet: '{target}'")
    ws = wb[target]
    p(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    p()

    p("=== ALL NON-EMPTY CELLS ===")
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                p(f"  R{cell.row}C{cell.column} ({cell.coordinate}): {repr(cell.value)}")

    p()
    p("=== ROW-BY-ROW DUMP ===")
    for row in ws.iter_rows(values_only=True):
        if any(v is not None for v in row):
            p(str(row))

except Exception as e:
    p("ERROR: " + str(e))
    import traceback
    p(traceback.format_exc())
finally:
    out.close()
